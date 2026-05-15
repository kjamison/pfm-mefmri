#!/usr/bin/env python3
"""Assign canonical network identities to fixed PFM Infomap communities.

This post-processing step does not alter Infomap or force a fixed community
count. It labels each subject-specific Infomap community using canonical FC and
spatial priors, reports ambiguous communities for review, and writes consensus
maps across density columns.
"""

from __future__ import annotations

import argparse
import csv
import shutil
import subprocess
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import nibabel as nib
import numpy as np
import scipy.io as sio
from nibabel.cifti2.cifti2_axes import ScalarAxis

EPS = 1e-8


def bool_int(value: str) -> int:
    ivalue = int(value)
    if ivalue not in (0, 1):
        raise argparse.ArgumentTypeError("expected 0 or 1")
    return ivalue


def zscore_vec(x: np.ndarray) -> np.ndarray:
    x = np.asarray(x, dtype=np.float64)
    mu = np.nanmean(x)
    sd = np.nanstd(x)
    if not np.isfinite(sd) or sd < EPS:
        return np.zeros_like(x, dtype=np.float64)
    return (x - mu) / sd


def zscore_rows(x: np.ndarray) -> np.ndarray:
    x = np.asarray(x, dtype=np.float64)
    mu = np.nanmean(x, axis=1, keepdims=True)
    sd = np.nanstd(x, axis=1, keepdims=True)
    sd = np.where(np.isfinite(sd) & (sd > EPS), sd, 1.0)
    return np.nan_to_num((x - mu) / sd)


def corr_columns_against_vector(cols: np.ndarray, vec: np.ndarray) -> np.ndarray:
    v = zscore_vec(vec)
    c = np.asarray(cols, dtype=np.float64)
    c = c - np.nanmean(c, axis=0, keepdims=True)
    csd = np.nanstd(c, axis=0, keepdims=True)
    csd = np.where(np.isfinite(csd) & (csd > EPS), csd, 1.0)
    c = np.nan_to_num(c / csd)
    return np.nan_to_num((v[:, None] * c).mean(axis=0))


def softmax(row: np.ndarray) -> np.ndarray:
    x = np.asarray(row, dtype=np.float64)
    x = np.nan_to_num(x, nan=-1e9, posinf=1e9, neginf=-1e9)
    x = x - np.max(x)
    ex = np.exp(x)
    den = ex.sum()
    if den <= EPS:
        return np.ones_like(ex) / max(ex.size, 1)
    return ex / den


def mat_string(x) -> str:
    arr = np.asarray(x)
    if arr.size == 1:
        x = arr.item()
    if isinstance(x, bytes):
        return x.decode("utf-8", errors="ignore").strip()
    if isinstance(x, np.ndarray):
        vals = x.ravel()
        if vals.dtype.kind in ("U", "S"):
            return "".join(str(v) for v in vals).strip()
        if vals.size == 1:
            return mat_string(vals[0])
    s = str(x).strip()
    return s if s else "Network"


def load_priors(path: Path) -> Tuple[np.ndarray, np.ndarray, List[str], np.ndarray]:
    mat = sio.loadmat(path, squeeze_me=True, struct_as_record=False)
    if "Priors" not in mat:
        raise ValueError(f"{path} does not contain a Priors struct")
    pri = mat["Priors"]
    if not hasattr(pri, "FC") or not hasattr(pri, "Spatial"):
        raise ValueError("Priors.mat must contain Priors.FC and Priors.Spatial")
    pri_fc = np.asarray(pri.FC, dtype=np.float64)
    pri_spatial = np.asarray(pri.Spatial, dtype=np.float64)
    if pri_fc.ndim != 2 or pri_spatial.ndim != 2:
        raise ValueError("Priors.FC and Priors.Spatial must be 2D")
    n_net = pri_fc.shape[1]
    if pri_spatial.shape[1] != n_net:
        raise ValueError("Priors.FC and Priors.Spatial must have the same number of networks")

    if hasattr(pri, "NetworkLabels"):
        raw = np.asarray(pri.NetworkLabels, dtype=object).ravel()
        labels = [mat_string(x) for x in raw[:n_net]]
    else:
        labels = []
    labels += [f"Network {i:02d}" for i in range(len(labels) + 1, n_net + 1)]

    if hasattr(pri, "NetworkColors"):
        colors = np.asarray(pri.NetworkColors, dtype=np.float64)
        colors = np.atleast_2d(colors)[:n_net, :3]
        if colors.size and float(np.nanmax(colors)) > 1.0:
            colors = colors / 255.0
    else:
        colors = np.zeros((0, 3), dtype=np.float64)
    if colors.shape[0] < n_net:
        rng = np.random.default_rng(7)
        pad = rng.uniform(0.25, 0.85, size=(n_net - colors.shape[0], 3))
        colors = np.vstack([colors, pad])
    colors = np.clip(np.nan_to_num(colors[:n_net, :3], nan=0.55), 0.0, 1.0)
    return pri_fc, pri_spatial, labels[:n_net], colors


def cortical_grayordinates(axis) -> np.ndarray:
    idx: List[int] = []
    for name, slc, _ in axis.iter_structures():
        if name in ("CIFTI_STRUCTURE_CORTEX_LEFT", "CIFTI_STRUCTURE_CORTEX_RIGHT"):
            stop = slc.stop if slc.stop is not None else axis.size
            idx.extend(range(slc.start, stop))
    return np.asarray(idx, dtype=np.int64)


def save_dscalar(ref_img: nib.Cifti2Image, maps: np.ndarray, names: Sequence[str], out_path: Path) -> None:
    arr = np.asarray(maps, dtype=np.float32)
    if arr.ndim == 1:
        arr = arr[None, :]
    axis = ref_img.header.get_axis(1)
    scalar = ScalarAxis(list(names))
    hdr = nib.Cifti2Header.from_axes((scalar, axis))
    nib.save(nib.Cifti2Image(arr, hdr, nifti_header=ref_img.nifti_header), str(out_path))


def write_label_list(path: Path, labels: Sequence[str], colors: np.ndarray, unassigned_value: int) -> None:
    with path.open("w", encoding="utf-8") as f:
        for i, label in enumerate(labels, start=1):
            rgb = np.clip(np.rint(colors[i - 1] * 255.0), 0, 255).astype(int)
            f.write(f"{label}\n")
            f.write(f"{i} {rgb[0]} {rgb[1]} {rgb[2]} 255\n")
        f.write("Unassigned\n")
        f.write(f"{int(unassigned_value)} 128 128 128 255\n")


def wb_available(wb_command: str) -> bool:
    return bool(wb_command) and shutil.which(wb_command) is not None


def import_dlabel(
    wb_command: str,
    ref_img: nib.Cifti2Image,
    values: np.ndarray,
    label_list: Path,
    out_path: Path,
    tmp_prefix: Path,
) -> bool:
    tmp = tmp_prefix.with_suffix(".dscalar.nii")
    save_dscalar(ref_img, values, ["labels"], tmp)
    if not wb_available(wb_command):
        fallback = out_path.with_suffix("").with_suffix(".dscalar.nii")
        tmp.replace(fallback)
        print(f"[infomap_labeler] WARNING: {wb_command} unavailable; wrote scalar fallback {fallback}")
        return False
    subprocess.run(
        [wb_command, "-cifti-label-import", str(tmp), str(label_list), str(out_path), "-discard-others"],
        check=True,
    )
    try:
        tmp.unlink()
    except OSError:
        pass
    return True


def write_borders(wb_command: str, dlabel: Path, left_surf: str, right_surf: str, out_prefix: Path) -> None:
    if not wb_available(wb_command) or not dlabel.exists():
        return
    if left_surf:
        ls = Path(left_surf)
        if ls.exists():
            subprocess.run([wb_command, "-cifti-label-to-border", str(dlabel), "-border", str(ls), str(out_prefix) + ".L.border"], check=True)
    if right_surf:
        rs = Path(right_surf)
        if rs.exists():
            subprocess.run([wb_command, "-cifti-label-to-border", str(dlabel), "-border", str(rs), str(out_prefix) + ".R.border"], check=True)


def write_optional_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "ManualCorrections"
    with csv_path.open(newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    if ws.max_row:
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 48)
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(xlsx_path)


def write_manual_correction_sheet(
    outdir: Path,
    prefix: str,
    rows: Sequence[Dict[str, object]],
    labels: Sequence[str],
    unassigned_value: int,
) -> None:
    fields = [
        "density_index",
        "community_id",
        "community_size",
        "current_label_index",
        "current_label_name",
        "manual_label_index",
        "manual_label_name",
        "review_flag",
        "review_reason",
        "rank1_label_name",
        "rank1_probability",
        "rank2_label_name",
        "rank2_probability",
        "confidence_margin",
        "fc_score_assigned",
        "spatial_score_assigned",
        "fc_top_label_name",
        "spatial_top_label_name",
        "fc_spatial_agree",
        "notes",
    ]
    csv_path = outdir / f"{prefix}_ManualCorrections.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "density_index": row["density_index"],
                    "community_id": row["community_id"],
                    "community_size": row["community_size"],
                    "current_label_index": row["assigned_label_index"],
                    "current_label_name": row["assigned_label_name"],
                    "manual_label_index": "",
                    "manual_label_name": "",
                    "review_flag": row["review_flag"],
                    "review_reason": row["review_reason"],
                    "rank1_label_name": row["rank1_label_name"],
                    "rank1_probability": row["rank1_probability"],
                    "rank2_label_name": row["rank2_label_name"],
                    "rank2_probability": row["rank2_probability"],
                    "confidence_margin": row["confidence_margin"],
                    "fc_score_assigned": row["fc_score_assigned"],
                    "spatial_score_assigned": row["spatial_score_assigned"],
                    "fc_top_label_name": row["fc_top_label_name"],
                    "spatial_top_label_name": row["spatial_top_label_name"],
                    "fc_spatial_agree": row["fc_spatial_agree"],
                    "notes": "",
                }
            )
    label_key = outdir / f"{prefix}_ManualCorrections_LabelKey.csv"
    with label_key.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["label_index", "label_name"])
        for i, label in enumerate(labels, 1):
            writer.writerow([i, label])
        writer.writerow([int(unassigned_value), "Unassigned"])
    write_optional_xlsx(csv_path, outdir / f"{prefix}_ManualCorrections.xlsx")
    print(f"[infomap_labeler] wrote {csv_path}")
    print(f"[infomap_labeler] wrote {label_key}")


def skewness_nonzero(x: np.ndarray) -> float:
    vals = np.asarray(x, dtype=np.float64)
    vals = vals[np.isfinite(vals) & (vals != 0)]
    if vals.size < 3:
        return 0.0
    z = zscore_vec(vals)
    return float(np.mean(z ** 3))


def adaptive_weights(
    pri_spatial: np.ndarray,
    lower: float,
    upper: float,
    enabled: bool,
    direction: str = "spatial_from_skew",
) -> Tuple[np.ndarray, np.ndarray]:
    n_net = pri_spatial.shape[1]
    if not enabled:
        return np.ones(n_net, dtype=np.float64), np.ones(n_net, dtype=np.float64)
    raw = np.array([skewness_nonzero(pri_spatial[:, k]) for k in range(n_net)], dtype=np.float64)
    raw = np.nan_to_num(raw)
    lo, hi = float(np.min(raw)), float(np.max(raw))
    if abs(hi - lo) < EPS:
        scaled = np.full(n_net, (lower + upper) / 2.0, dtype=np.float64)
    else:
        scaled = lower + (raw - lo) / (hi - lo) * (upper - lower)
    scaled = np.clip(scaled, lower, upper)
    if direction == "fc_from_skew":
        fcw = scaled
    elif direction == "spatial_from_skew":
        fcw = 1.0 - scaled
    else:
        raise ValueError(f"Unknown adaptive weighting direction: {direction}")
    return fcw, 1.0 - fcw


def select_density_indices(n_cols: int, density_index: int) -> List[int]:
    if density_index == -1:
        return list(range(n_cols))
    if density_index <= 0:
        raise ValueError("--density-index must be -1 for all columns or a 1-based positive column index")
    idx = density_index - 1
    if idx >= n_cols:
        raise ValueError(f"--density-index {density_index} exceeds community columns ({n_cols})")
    return [idx]


def build_scores(
    community_ids: np.ndarray,
    communities: np.ndarray,
    data_tg: np.ndarray,
    cortex_idx: np.ndarray,
    cort_lookup: np.ndarray,
    pri_fc: np.ndarray,
    pri_spatial: np.ndarray,
    fc_prior_weight: np.ndarray,
    spatial_prior_weight: np.ndarray,
    args,
) -> Dict[str, np.ndarray]:
    n_comm = community_ids.size
    n_net = pri_fc.shape[1]
    n_time = data_tg.shape[0]
    cortex_ts = zscore_rows(data_tg[:, cortex_idx].T)
    pri_fc_z = zscore_rows(pri_fc.T).T

    fc_sim = np.zeros((n_comm, n_net), dtype=np.float64)
    spatial_score = np.zeros((n_comm, n_net), dtype=np.float64)
    comm_ts = np.zeros((n_comm, n_time), dtype=np.float64)
    sizes = np.zeros(n_comm, dtype=np.int64)
    cortical_sizes = np.zeros(n_comm, dtype=np.int64)

    for i, cid in enumerate(community_ids):
        members = np.where(communities == cid)[0]
        sizes[i] = members.size
        if members.size == 0:
            continue
        ts = np.nanmean(data_tg[:, members], axis=1)
        comm_ts[i, :] = zscore_vec(ts)
        fc_profile = (cortex_ts @ comm_ts[i, :]) / max(n_time - 1, 1)
        fc_sim[i, :] = corr_columns_against_vector(pri_fc_z, fc_profile)

        cmask = cort_lookup[members] >= 0
        cort_members = members[cmask]
        cortical_sizes[i] = cort_members.size
        if cort_members.size:
            spatial_score[i, :] = np.nanmean(pri_spatial[cort_lookup[cort_members], :], axis=0)

    fc_z = np.vstack([zscore_vec(row) for row in fc_sim]) if n_comm else np.zeros((0, n_net))
    if args.score_mode == "beta_product":
        score = (
            fc_sim * float(args.fc_weight) * fc_prior_weight[None, :]
        ) * (
            spatial_score * float(args.spatial_weight) * spatial_prior_weight[None, :]
        )
    else:
        score = (
            float(args.fc_weight) * fc_prior_weight[None, :] * fc_z
            + float(args.spatial_weight) * spatial_prior_weight[None, :] * np.log(np.maximum(spatial_score, EPS))
        )
    prob = np.vstack([softmax(row) for row in score]) if n_comm else np.zeros((0, n_net))
    return {
        "fc_sim": fc_sim,
        "spatial_score": spatial_score,
        "score": score,
        "prob": prob,
        "comm_ts": comm_ts,
        "sizes": sizes,
        "cortical_sizes": cortical_sizes,
    }


def optimize_labels(
    labels0: np.ndarray,
    score: np.ndarray,
    comm_ts: np.ndarray,
    pairwise_weight: float,
    threshold: float,
    max_iter: int,
    seed: int,
) -> Tuple[np.ndarray, List[Dict[str, object]]]:
    if pairwise_weight <= 0 or labels0.size <= 1:
        return labels0.copy()
    labels = labels0.copy()
    n_comm, n_net = score.shape
    corr = np.corrcoef(comm_ts)
    corr = np.nan_to_num(corr, nan=0.0)
    np.fill_diagonal(corr, 0.0)
    adj = np.where(corr >= float(threshold), corr, 0.0)
    rng = np.random.default_rng(int(seed))
    for _ in range(max(0, int(max_iter))):
        changed = 0
        for c in rng.permutation(n_comm):
            if labels[c] <= 0:
                continue
            costs = -score[c, :].copy()
            nb = adj[c, :] > 0
            if np.any(nb):
                nb_labels = labels[nb]
                nb_w = adj[c, nb]
                for k in range(1, n_net + 1):
                    costs[k - 1] += pairwise_weight * float(nb_w[nb_labels != k].sum())
            new_label = int(np.argmin(costs)) + 1
            if new_label != int(labels[c]):
                labels[c] = new_label
                changed += 1
        if changed == 0:
            break
    return labels


def anchor_context_labels(
    init_labels: np.ndarray,
    score: np.ndarray,
    prob: np.ndarray,
    fc_sim: np.ndarray,
    spatial_score: np.ndarray,
    comm_ts: np.ndarray,
    sizes: np.ndarray,
    args,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    labels = init_labels.copy()
    n_comm, n_net = score.shape
    if not bool(int(args.context_enable)) or n_comm <= 1:
        return labels, np.zeros(n_comm, dtype=np.int8), np.zeros(n_comm, dtype=np.int64), np.zeros(n_comm, dtype=np.float64)

    rank = np.argsort(score, axis=1)[:, ::-1]
    r1 = rank[:, 0]
    r2 = rank[:, 1] if n_net > 1 else rank[:, 0]
    confidence = prob[np.arange(n_comm), r1] - prob[np.arange(n_comm), r2]
    fc_top = np.argmax(fc_sim, axis=1) + 1
    spatial_top = np.argmax(spatial_score, axis=1) + 1
    fc_assigned = fc_sim[np.arange(n_comm), np.maximum(labels - 1, 0)]

    anchors = (
        (confidence >= float(args.context_min_anchor_confidence))
        & (fc_assigned >= float(args.context_min_anchor_fc_similarity))
        & (sizes >= int(args.min_community_size))
    )
    if int(args.context_anchor_require_fc_spatial_agree) == 1:
        anchors &= fc_top == spatial_top
    if not np.any(anchors):
        return labels, np.zeros(n_comm, dtype=np.int8), np.zeros(n_comm, dtype=np.int64), np.zeros(n_comm, dtype=np.float64)

    target = np.ones(n_comm, dtype=bool)
    if int(args.context_only_ambiguous) == 1:
        target = confidence < float(args.confidence_threshold)
    target &= ~anchors

    corr = np.corrcoef(comm_ts)
    corr = np.nan_to_num(corr, nan=0.0)
    np.fill_diagonal(corr, 0.0)
    changed = np.zeros(n_comm, dtype=np.int8)
    context_top = np.zeros(n_comm, dtype=np.int64)
    context_support = np.zeros(n_comm, dtype=np.float64)

    for c in np.where(target)[0]:
        linked = anchors & (corr[c, :] >= float(args.context_min_community_fc))
        if not np.any(linked):
            continue
        support = np.zeros(n_net, dtype=np.float64)
        for lab in range(1, n_net + 1):
            support[lab - 1] = float(corr[c, linked & (labels == lab)].sum())
        if np.max(support) <= EPS:
            continue
        support_z = zscore_vec(support)
        context_score = score[c, :] + float(args.context_weight) * support_z
        if float(args.context_min_spatial_score) > 0:
            context_score = context_score.copy()
            context_score[spatial_score[c, :] < float(args.context_min_spatial_score)] = -1e9
        new_label = int(np.argmax(context_score)) + 1
        old_label = int(labels[c])
        context_top[c] = new_label
        context_support[c] = float(support[new_label - 1])
        if new_label != old_label:
            old_score = float(context_score[old_label - 1]) if 1 <= old_label <= n_net else -1e9
            new_score = float(context_score[new_label - 1])
            if new_score - old_score >= float(args.context_min_switch_margin):
                labels[c] = new_label
                changed[c] = 1
    return labels, changed, context_top, context_support


def review_flags(
    size: int,
    confidence: float,
    fc_assigned: float,
    fc_top: int,
    spatial_top: int,
    changed: bool,
    assigned: int,
    args,
) -> Tuple[int, str]:
    reasons: List[str] = []
    if size < int(args.min_community_size):
        reasons.append("small_community")
    if confidence < float(args.confidence_threshold):
        reasons.append("low_confidence")
    if fc_assigned < float(args.min_fc_similarity):
        reasons.append("low_fc_similarity")
    if fc_top != spatial_top:
        reasons.append("fc_spatial_disagree")
    if changed:
        reasons.append("optimizer_changed_label")
    if assigned == int(args.unassigned_value):
        reasons.append("unassigned")
    return (1 if reasons else 0), ";".join(reasons)


def density_label(
    density_col: int,
    community_ids: np.ndarray,
    communities: np.ndarray,
    data_tg: np.ndarray,
    ref_img: nib.Cifti2Image,
    cortex_idx: np.ndarray,
    cort_lookup: np.ndarray,
    pri_fc: np.ndarray,
    pri_spatial: np.ndarray,
    labels: Sequence[str],
    colors: np.ndarray,
    fc_prior_weight: np.ndarray,
    spatial_prior_weight: np.ndarray,
    label_list: Path,
    args,
) -> np.ndarray:
    outdir = Path(args.outdir)
    prefix = str(args.outfile_prefix)
    n_gray = communities.size
    n_net = len(labels)
    dens_tag = f"Density{density_col + 1:02d}"
    label_map = np.zeros(n_gray, dtype=np.float32)
    conf_map = np.zeros(n_gray, dtype=np.float32)

    metrics = build_scores(
        community_ids,
        communities,
        data_tg,
        cortex_idx,
        cort_lookup,
        pri_fc,
        pri_spatial,
        fc_prior_weight,
        spatial_prior_weight,
        args,
    )
    score = metrics["score"]
    prob = metrics["prob"]
    fc_sim = metrics["fc_sim"]
    spatial_score = metrics["spatial_score"]
    sizes = metrics["sizes"]

    rank = np.argsort(score, axis=1)[:, ::-1] if community_ids.size else np.zeros((0, n_net), dtype=np.int64)
    init_labels = rank[:, 0] + 1 if community_ids.size else np.zeros(0, dtype=np.int64)
    opt_labels = optimize_labels(
        init_labels,
        score,
        metrics["comm_ts"],
        float(args.pairwise_weight),
        float(args.pairwise_neighbor_threshold),
        int(args.max_optimizer_iterations),
        int(args.optimizer_random_seed),
    )
    context_labels, context_changed, context_top, context_support = anchor_context_labels(
        opt_labels,
        score,
        prob,
        fc_sim,
        spatial_score,
        metrics["comm_ts"],
        sizes,
        args,
    )

    rows = []
    ambiguous = []
    for i, cid in enumerate(community_ids):
        r1 = int(rank[i, 0])
        r2 = int(rank[i, 1]) if n_net > 1 else r1
        top_prob = float(prob[i, r1])
        second_prob = float(prob[i, r2]) if n_net > 1 else 0.0
        confidence = top_prob - second_prob
        beta_conf = float((score[i, r1] - score[i, r2]) / (abs(score[i, r2]) + EPS)) if n_net > 1 else float("inf")
        threshold_confidence = (
            beta_conf
            if args.score_mode == "beta_product" and args.beta_product_threshold_mode == "beta_ratio"
            else confidence
        )
        assigned = int(context_labels[i])
        fc_top = int(np.argmax(fc_sim[i, :])) + 1
        spatial_top = int(np.argmax(spatial_score[i, :])) + 1
        changed = assigned != int(init_labels[i])
        fc_assigned = float(fc_sim[i, assigned - 1])
        strict_fail = (
            sizes[i] < int(args.min_community_size)
            or threshold_confidence < float(args.confidence_threshold)
            or fc_assigned < float(args.min_fc_similarity)
        )
        if int(args.strict_thresholding) == 1 and strict_fail:
            assigned = int(args.unassigned_value)
        flag, reason = review_flags(
            int(sizes[i]),
            threshold_confidence,
            fc_assigned,
            fc_top,
            spatial_top,
            changed,
            assigned,
            args,
        )
        members = communities == cid
        label_map[members] = float(assigned)
        conf_map[members] = float(confidence)
        row = {
            "density_index": density_col + 1,
            "community_id": int(cid),
            "community_size": int(sizes[i]),
            "assigned_label_index": int(assigned),
            "assigned_label_name": "Unassigned" if assigned == int(args.unassigned_value) else labels[assigned - 1],
            "assigned_label_after_optimizer": int(opt_labels[i]),
            "rank1_label_name": labels[r1],
            "rank1_score": float(score[i, r1]),
            "rank1_probability": top_prob,
            "rank2_label_name": labels[r2],
            "rank2_score": float(score[i, r2]),
            "rank2_probability": second_prob,
            "confidence_margin": confidence,
            "beta_confidence_ratio": beta_conf,
            "fc_score_assigned": fc_assigned,
            "spatial_score_assigned": float(spatial_score[i, assigned - 1]) if 1 <= assigned <= n_net else 0.0,
            "fc_top_label_name": labels[fc_top - 1],
            "spatial_top_label_name": labels[spatial_top - 1],
            "fc_spatial_agree": int(fc_top == spatial_top),
            "adaptive_fc_weight_assigned": float(fc_prior_weight[assigned - 1]) if 1 <= assigned <= n_net else 0.0,
            "adaptive_spatial_weight_assigned": float(spatial_prior_weight[assigned - 1]) if 1 <= assigned <= n_net else 0.0,
            "review_flag": flag,
            "review_reason": reason,
            "optimizer_changed_label": int(changed),
            "context_changed_label": int(context_changed[i]),
            "context_top_label_name": labels[int(context_top[i]) - 1] if 1 <= int(context_top[i]) <= n_net else "",
            "context_support_assigned": float(context_support[i]),
        }
        rows.append(row)
        if flag:
            ambiguous.append(row)

    table_fields = [
        "density_index",
        "community_id",
        "community_size",
        "assigned_label_index",
        "assigned_label_name",
        "assigned_label_after_optimizer",
        "rank1_label_name",
        "rank1_score",
        "rank1_probability",
        "rank2_label_name",
        "rank2_score",
        "rank2_probability",
        "confidence_margin",
        "beta_confidence_ratio",
        "fc_score_assigned",
        "spatial_score_assigned",
        "fc_top_label_name",
        "spatial_top_label_name",
        "fc_spatial_agree",
        "adaptive_fc_weight_assigned",
        "adaptive_spatial_weight_assigned",
        "review_flag",
        "review_reason",
        "optimizer_changed_label",
        "context_changed_label",
        "context_top_label_name",
        "context_support_assigned",
    ]
    for suffix, data_rows in (("CommunityTable", rows), ("AmbiguousCommunities", ambiguous)):
        path = outdir / f"{prefix}_{dens_tag}_{suffix}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=table_fields)
            writer.writeheader()
            writer.writerows(data_rows)
        print(f"[infomap_labeler] wrote {path}")

    dlabel = outdir / f"{prefix}_{dens_tag}.dlabel.nii"
    wrote_dlabel = import_dlabel(args.wb_command, ref_img, label_map, label_list, dlabel, outdir / f"Tmp_{prefix}_{dens_tag}")
    save_dscalar(ref_img, conf_map, ["confidence"], outdir / f"{prefix}_{dens_tag}_Confidence.dscalar.nii")
    if wrote_dlabel:
        print(f"[infomap_labeler] wrote {dlabel}")
    return label_map, rows


def mode_consensus(label_maps: np.ndarray, n_net: int, unassigned: int) -> np.ndarray:
    n_density, n_gray = label_maps.shape
    out = np.zeros(n_gray, dtype=np.float32)
    for v in range(n_gray):
        vals = label_maps[:, v].astype(np.int64)
        vals = vals[vals > 0]
        non_unassigned = vals[vals != int(unassigned)]
        use = non_unassigned if non_unassigned.size else vals
        if use.size == 0:
            continue
        counts = np.bincount(use, minlength=max(n_net, int(unassigned)) + 1)
        max_count = counts.max()
        tied = np.where(counts == max_count)[0]
        # Deterministic tie break: use latest/highest-density column among tied labels.
        chosen = int(tied[0])
        for lab in vals[::-1]:
            if int(lab) in tied:
                chosen = int(lab)
                break
        out[v] = float(chosen)
    return out


def probability_consensus(label_maps: np.ndarray, n_net: int, unassigned: int) -> np.ndarray:
    out = np.zeros((n_net, label_maps.shape[1]), dtype=np.float32)
    valid = (label_maps > 0) & (label_maps != int(unassigned))
    denom = valid.sum(axis=0).astype(np.float32)
    denom[denom == 0] = 1.0
    for k in range(1, n_net + 1):
        out[k - 1, :] = (label_maps == k).sum(axis=0) / denom
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Canonical network labeling for PFM Infomap communities")
    ap.add_argument("--in-cifti", required=True)
    ap.add_argument("--communities-cifti", required=True)
    ap.add_argument("--priors-mat", required=True)
    ap.add_argument("--outdir", required=True)
    ap.add_argument("--outfile-prefix", default="InfomapNetworkLabels")
    ap.add_argument("--density-index", type=int, default=-1)
    ap.add_argument("--score-mode", choices=("additive", "beta_product"), default="additive")
    ap.add_argument(
        "--beta-product-threshold-mode",
        choices=("beta_ratio", "softmax_margin"),
        default="beta_ratio",
        help="Confidence metric used for thresholding/review flags in beta_product mode.",
    )
    ap.add_argument("--adaptive-prior-weighting", type=bool_int, default=0)
    ap.add_argument("--fc-weight", type=float, default=1.0)
    ap.add_argument("--spatial-weight", type=float, default=1.0)
    ap.add_argument("--pairwise-weight", type=float, default=0.0)
    ap.add_argument("--confidence-threshold", type=float, default=0.15)
    ap.add_argument("--min-fc-similarity", type=float, default=0.33)
    ap.add_argument("--min-community-size", type=int, default=10)
    ap.add_argument("--unassigned-value", type=int, default=21)
    ap.add_argument("--strict-thresholding", type=bool_int, default=0)
    ap.add_argument("--adaptive-fc-lower-bound", type=float, default=0.25)
    ap.add_argument("--adaptive-fc-upper-bound", type=float, default=0.75)
    ap.add_argument(
        "--adaptive-weight-direction",
        choices=("spatial_from_skew", "fc_from_skew"),
        default="spatial_from_skew",
        help=(
            "How spatial-prior skewness changes adaptive weights. "
            "spatial_from_skew gives spatially stereotyped networks more spatial weight; "
            "fc_from_skew preserves the older beta-style direction."
        ),
    )
    ap.add_argument("--pairwise-neighbor-threshold", type=float, default=0.2)
    ap.add_argument("--max-optimizer-iterations", type=int, default=25)
    ap.add_argument("--optimizer-random-seed", type=int, default=44)
    ap.add_argument("--context-enable", type=bool_int, default=0)
    ap.add_argument("--context-mode", choices=("anchor_propagation",), default="anchor_propagation")
    ap.add_argument("--context-weight", type=float, default=0.25)
    ap.add_argument("--context-min-anchor-confidence", type=float, default=0.25)
    ap.add_argument("--context-min-anchor-fc-similarity", type=float, default=0.40)
    ap.add_argument("--context-anchor-require-fc-spatial-agree", type=bool_int, default=1)
    ap.add_argument("--context-min-community-fc", type=float, default=0.25)
    ap.add_argument("--context-only-ambiguous", type=bool_int, default=1)
    ap.add_argument("--context-min-spatial-score", type=float, default=0.01)
    ap.add_argument("--context-min-switch-margin", type=float, default=0.05)
    ap.add_argument("--left-surf", default="")
    ap.add_argument("--right-surf", default="")
    ap.add_argument("--wb-command", default="wb_command")
    args = ap.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    img = nib.load(args.in_cifti)
    comm_img = nib.load(args.communities_cifti)
    data = np.asarray(img.get_fdata(dtype=np.float32))
    comm_data = np.asarray(comm_img.get_fdata(dtype=np.float32))
    if data.ndim != 2 or comm_data.ndim != 2:
        raise ValueError("Input CIFTI and communities CIFTI must be 2D")
    if data.shape[1] != comm_data.shape[1]:
        raise ValueError(f"Grayordinate mismatch: data={data.shape[1]} communities={comm_data.shape[1]}")

    axis = img.header.get_axis(1)
    cortex_idx = cortical_grayordinates(axis)
    if cortex_idx.size == 0:
        raise ValueError("No cortical grayordinates found in input CIFTI")
    pri_fc, pri_spatial, labels, colors = load_priors(Path(args.priors_mat))
    n_cort = cortex_idx.size
    if pri_fc.shape[0] < n_cort or pri_spatial.shape[0] < n_cort:
        raise ValueError(
            f"Priors rows must cover cortical grayordinates ({n_cort}); got FC={pri_fc.shape[0]} Spatial={pri_spatial.shape[0]}"
        )
    pri_fc = np.nan_to_num(pri_fc[:n_cort, :])
    pri_spatial = np.clip(np.nan_to_num(pri_spatial[:n_cort, :]), 0.0, None)
    n_net = pri_fc.shape[1]
    if 1 <= int(args.unassigned_value) <= n_net:
        raise ValueError(
            f"--unassigned-value ({args.unassigned_value}) overlaps canonical network labels 1..{n_net}; "
            "choose a value outside that range"
        )

    fc_prior_weight, spatial_prior_weight = adaptive_weights(
        pri_spatial,
        float(args.adaptive_fc_lower_bound),
        float(args.adaptive_fc_upper_bound),
        bool(args.adaptive_prior_weighting),
        args.adaptive_weight_direction,
    )

    label_list = outdir / f"{args.outfile_prefix}_LabelListFile.txt"
    write_label_list(label_list, labels, colors, int(args.unassigned_value))

    cort_lookup = np.full(axis.size, -1, dtype=np.int64)
    cort_lookup[cortex_idx] = np.arange(cortex_idx.size, dtype=np.int64)
    density_indices = select_density_indices(comm_data.shape[0], int(args.density_index))
    label_maps: List[np.ndarray] = []
    manual_rows: List[Dict[str, object]] = []

    for di in density_indices:
        communities = np.rint(comm_data[di, :]).astype(np.int64)
        community_ids = np.unique(communities[communities > 0])
        print(f"[infomap_labeler] density={di + 1} communities={community_ids.size}")
        if community_ids.size == 0:
            label_maps.append(np.zeros(axis.size, dtype=np.float32))
            continue
        label_map, rows = density_label(
            di,
            community_ids,
            communities,
            data,
            img,
            cortex_idx,
            cort_lookup,
            pri_fc,
            pri_spatial,
            labels,
            colors,
            fc_prior_weight,
            spatial_prior_weight,
            label_list,
            args,
        )
        label_maps.append(label_map)
        manual_rows.extend(rows)

    labels_arr = np.vstack(label_maps).astype(np.float32)
    mode = mode_consensus(labels_arr, n_net, int(args.unassigned_value))
    prob = probability_consensus(labels_arr, n_net, int(args.unassigned_value))

    consensus = outdir / f"{args.outfile_prefix}_ModeConsensus.dlabel.nii"
    wrote_consensus_dlabel = import_dlabel(args.wb_command, img, mode, label_list, consensus, outdir / f"Tmp_{args.outfile_prefix}_ModeConsensus")
    save_dscalar(img, prob, labels, outdir / f"{args.outfile_prefix}_ProbabilityConsensus.dscalar.nii")
    write_borders(
        args.wb_command,
        consensus,
        args.left_surf,
        args.right_surf,
        outdir / f"{args.outfile_prefix}_Consensus",
    )
    if wrote_consensus_dlabel:
        print(f"[infomap_labeler] wrote {consensus}")
    print(f"[infomap_labeler] wrote {outdir / f'{args.outfile_prefix}_ProbabilityConsensus.dscalar.nii'}")
    write_manual_correction_sheet(outdir, str(args.outfile_prefix), manual_rows, labels, int(args.unassigned_value))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
