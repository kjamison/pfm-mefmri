#!/usr/bin/env python3
"""Apply manual corrections to PFM Infomap network labels.

This module reads the community-level manual correction sheet emitted by
pfm_infomap_labeler.py, applies any non-empty manual labels to fixed Infomap
communities, and writes adjusted dense labels plus consensus maps. It does not
rerun Infomap or recompute FC/spatial evidence.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import nibabel as nib
import numpy as np

from pfm_infomap_labeler import (
    import_dlabel,
    load_priors,
    mode_consensus,
    probability_consensus,
    save_dscalar,
    write_borders,
    write_label_list,
)


def read_rows(path: Path) -> List[Dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("Reading .xlsx manual corrections requires openpyxl") from exc
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        fields = [str(v or "").strip() for v in rows[0]]
        out: List[Dict[str, str]] = []
        for vals in rows[1:]:
            row = {fields[i]: "" if vals[i] is None else str(vals[i]).strip() for i in range(min(len(fields), len(vals)))}
            if any(row.values()):
                out.append(row)
        return out
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(f)]


def label_lookup(labels: Sequence[str], unassigned_value: int) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for i, label in enumerate(labels, 1):
        out[str(i)] = i
        out[label] = i
        out[label.lower()] = i
    out[str(int(unassigned_value))] = int(unassigned_value)
    out["Unassigned"] = int(unassigned_value)
    out["unassigned"] = int(unassigned_value)
    return out


def parse_label(row: Dict[str, str], labels: Sequence[str], unassigned_value: int) -> int:
    lookup = label_lookup(labels, unassigned_value)
    manual_idx = row.get("manual_label_index", "")
    manual_name = row.get("manual_label_name", "")
    current_idx = row.get("current_label_index", row.get("assigned_label_index", ""))
    current_name = row.get("current_label_name", row.get("assigned_label_name", ""))

    for val in (manual_idx, manual_name, current_idx, current_name):
        val = str(val).strip()
        if not val:
            continue
        if val in lookup:
            return lookup[val]
        if val.lower() in lookup:
            return lookup[val.lower()]
    raise ValueError(f"Could not resolve label for density={row.get('density_index')} community={row.get('community_id')}")


def current_label(row: Dict[str, str], labels: Sequence[str], unassigned_value: int) -> int:
    lookup = label_lookup(labels, unassigned_value)
    for key in ("current_label_index", "assigned_label_index", "current_label_name", "assigned_label_name"):
        val = str(row.get(key, "")).strip()
        if val in lookup:
            return lookup[val]
        if val.lower() in lookup:
            return lookup[val.lower()]
    return 0


def manual_value_present(row: Dict[str, str]) -> bool:
    return bool(str(row.get("manual_label_index", "")).strip() or str(row.get("manual_label_name", "")).strip())


def build_assignments(
    rows: Iterable[Dict[str, str]],
    labels: Sequence[str],
    unassigned_value: int,
) -> Tuple[Dict[int, Dict[int, int]], List[Dict[str, object]]]:
    assignments: Dict[int, Dict[int, int]] = {}
    applied: List[Dict[str, object]] = []
    for row in rows:
        if not row.get("density_index") or not row.get("community_id"):
            continue
        density = int(float(row["density_index"]))
        community = int(float(row["community_id"]))
        label = parse_label(row, labels, unassigned_value)
        assignments.setdefault(density, {})[community] = label
        old = current_label(row, labels, unassigned_value)
        changed = manual_value_present(row) and old != label
        applied.append(
            {
                "density_index": density,
                "community_id": community,
                "original_label_index": old,
                "corrected_label_index": label,
                "manual_override_present": int(manual_value_present(row)),
                "changed": int(changed),
                "notes": row.get("notes", ""),
            }
        )
    return assignments, applied


def write_applied_table(path: Path, rows: Sequence[Dict[str, object]], labels: Sequence[str], unassigned_value: int) -> None:
    fields = [
        "density_index",
        "community_id",
        "original_label_index",
        "original_label_name",
        "corrected_label_index",
        "corrected_label_name",
        "manual_override_present",
        "changed",
        "notes",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            old = int(row["original_label_index"])
            new = int(row["corrected_label_index"])
            out = dict(row)
            out["original_label_name"] = "Unassigned" if old == int(unassigned_value) else labels[old - 1] if 1 <= old <= len(labels) else ""
            out["corrected_label_name"] = "Unassigned" if new == int(unassigned_value) else labels[new - 1] if 1 <= new <= len(labels) else ""
            writer.writerow(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="Apply manual corrections to PFM Infomap community labels")
    ap.add_argument("--communities-cifti", required=True)
    ap.add_argument("--manual-corrections", required=True)
    ap.add_argument("--priors-mat", required=True)
    ap.add_argument("--outdir", required=True)
    ap.add_argument("--outfile-prefix", default="InfomapNetworkLabels_ManualAdjusted")
    ap.add_argument("--density-index", type=int, default=-1)
    ap.add_argument("--unassigned-value", type=int, default=21)
    ap.add_argument("--left-surf", default="")
    ap.add_argument("--right-surf", default="")
    ap.add_argument("--wb-command", default="wb_command")
    args = ap.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    comm_img = nib.load(args.communities_cifti)
    comm_data = np.asarray(comm_img.get_fdata(dtype=np.float32))
    if comm_data.ndim != 2:
        raise ValueError("Communities CIFTI must be 2D")

    _, _, labels, colors = load_priors(Path(args.priors_mat))
    n_net = len(labels)
    label_list = outdir / f"{args.outfile_prefix}_LabelListFile.txt"
    write_label_list(label_list, labels, colors, int(args.unassigned_value))

    rows = read_rows(Path(args.manual_corrections))
    assignments, applied = build_assignments(rows, labels, int(args.unassigned_value))
    if not assignments:
        raise ValueError("No manual correction rows were found")

    if int(args.density_index) == -1:
        density_indices = sorted(d - 1 for d in assignments.keys() if 1 <= d <= comm_data.shape[0])
    elif int(args.density_index) > 0:
        density_indices = [int(args.density_index) - 1]
    else:
        raise ValueError("--density-index must be -1 or a 1-based positive index")

    label_maps: List[np.ndarray] = []
    for di in density_indices:
        density = di + 1
        communities = np.rint(comm_data[di, :]).astype(np.int64)
        label_map = np.zeros(comm_data.shape[1], dtype=np.float32)
        for community, label in assignments.get(density, {}).items():
            label_map[communities == int(community)] = float(label)
        dens_tag = f"Density{density:02d}"
        dlabel = outdir / f"{args.outfile_prefix}_{dens_tag}.dlabel.nii"
        wrote = import_dlabel(
            args.wb_command,
            comm_img,
            label_map,
            label_list,
            dlabel,
            outdir / f"Tmp_{args.outfile_prefix}_{dens_tag}",
        )
        if wrote:
            print(f"[infomap_manual_labels] wrote {dlabel}")
        label_maps.append(label_map)

    labels_arr = np.vstack(label_maps).astype(np.float32)
    mode = mode_consensus(labels_arr, n_net, int(args.unassigned_value))
    prob = probability_consensus(labels_arr, n_net, int(args.unassigned_value))
    consensus = outdir / f"{args.outfile_prefix}_ModeConsensus.dlabel.nii"
    wrote_consensus = import_dlabel(
        args.wb_command,
        comm_img,
        mode,
        label_list,
        consensus,
        outdir / f"Tmp_{args.outfile_prefix}_ModeConsensus",
    )
    save_dscalar(comm_img, prob, labels, outdir / f"{args.outfile_prefix}_ProbabilityConsensus.dscalar.nii")
    write_borders(args.wb_command, consensus, args.left_surf, args.right_surf, outdir / f"{args.outfile_prefix}_Consensus")
    if wrote_consensus:
        print(f"[infomap_manual_labels] wrote {consensus}")
    print(f"[infomap_manual_labels] wrote {outdir / f'{args.outfile_prefix}_ProbabilityConsensus.dscalar.nii'}")
    applied_path = outdir / f"{args.outfile_prefix}_AppliedCorrections.csv"
    write_applied_table(applied_path, applied, labels, int(args.unassigned_value))
    print(f"[infomap_manual_labels] wrote {applied_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
